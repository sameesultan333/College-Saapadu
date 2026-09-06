"""
Daily Sales & GST report -> .xlsx

Money and quantities are written as real numbers (not formatted strings)
so a manager can pivot/sum them directly; formatting is applied via cell
number formats instead.
"""

from io import BytesIO
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

MONEY_FMT = "#,##0.00"
HEADER_FILL = PatternFill("solid", fgColor="EFEFEF")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(bottom=THIN)


def _fmt_date(iso: str) -> str:
    try:
        return date.fromisoformat(iso).strftime("%d %B %Y")
    except Exception:
        return iso


def _write_header(ws, headers, row=1):
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=title)
        c.font = BOLD
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center" if col > 1 else "left")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _money_cells(ws, first_row, last_row, cols):
    for r in range(first_row, last_row + 1):
        for c in cols:
            ws.cell(row=r, column=c).number_format = MONEY_FMT


def build_college_report_excel(report: dict) -> bytes:
    """College-wide workbook: canteen-wise summary plus combined detail."""
    wb = Workbook()
    g = report["grand_totals"]

    # ---------------- Canteen Summary ----------------
    ws = wb.active
    ws.title = "Canteen Summary"
    ws["A1"] = "COLLEGE SAAPAADU - Daily Sales & GST Report (All Canteens)"
    ws["A1"].font = Font(bold=True, size=13)

    meta = [
        ("College", report["college"]["name"]),
        ("Report Date", _fmt_date(report["report_date"])),
        ("Canteens", len(report["canteens"])),
        ("Generated At", report["generated_at"][:19].replace("T", " ") + " IST"),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v)

    header_row = 8
    _write_header(ws, ["Canteen", "Orders", "Items", "Gross Sales", "Taxable Value",
                       "CGST", "SGST", "Total GST"], row=header_row)
    r = header_row + 1
    for sec in report["canteens"]:
        t = sec["totals"]
        ws.cell(row=r, column=1, value=sec["canteen"]["name"])
        ws.cell(row=r, column=2, value=t["order_count"])
        ws.cell(row=r, column=3, value=t["item_count"])
        ws.cell(row=r, column=4, value=t["gross_sales"])
        ws.cell(row=r, column=5, value=t["taxable_sales"])
        ws.cell(row=r, column=6, value=t["cgst_amount"])
        ws.cell(row=r, column=7, value=t["sgst_amount"])
        ws.cell(row=r, column=8, value=t["total_gst"])
        r += 1

    ws.cell(row=r, column=1, value="TOTAL (ALL CANTEENS)")
    ws.cell(row=r, column=2, value=g["order_count"])
    ws.cell(row=r, column=3, value=g["item_count"])
    ws.cell(row=r, column=4, value=g["gross_sales"])
    ws.cell(row=r, column=5, value=g["taxable_sales"])
    ws.cell(row=r, column=6, value=g["cgst_amount"])
    ws.cell(row=r, column=7, value=g["sgst_amount"])
    ws.cell(row=r, column=8, value=g["total_gst"])
    for c in range(1, 9):
        ws.cell(row=r, column=c).font = BOLD
    _money_cells(ws, header_row + 1, r, [4, 5, 6, 7, 8])
    _autosize(ws, [30, 10, 10, 15, 16, 13, 13, 14])

    # ---------------- Item Summary (combined + per canteen) ----------------
    ws = wb.create_sheet("Item Summary")
    _write_header(ws, ["Scope", "Canteen", "Item", "Qty", "GST %", "Gross Sales",
                       "Taxable Value", "CGST", "SGST", "Total GST"])
    for it in report["combined_item_summary"]:
        ws.append(["All Canteens", "", it["name"], it["quantity"], it["gst_rate"],
                   it["gross_amount"], it["taxable_amount"], it["cgst_amount"],
                   it["sgst_amount"], it["total_gst_amount"]])
    for sec in report["canteens"]:
        for it in sec["item_summary"]:
            ws.append(["Canteen", sec["canteen"]["name"], it["name"], it["quantity"], it["gst_rate"],
                       it["gross_amount"], it["taxable_amount"], it["cgst_amount"],
                       it["sgst_amount"], it["total_gst_amount"]])
    if ws.max_row > 1:
        _money_cells(ws, 2, ws.max_row, [6, 7, 8, 9, 10])
    _autosize(ws, [13, 22, 26, 8, 8, 14, 15, 12, 12, 13])

    # ---------------- GST Summary ----------------
    ws = wb.create_sheet("GST Summary")
    _write_header(ws, ["Scope", "Canteen", "GST Rate %", "Gross", "Taxable Value",
                       "CGST", "SGST", "Total GST"])
    for gr in report["combined_gst_summary"]:
        ws.append(["All Canteens", "", gr["gst_rate"], gr["gross_amount"], gr["taxable_amount"],
                   gr["cgst_amount"], gr["sgst_amount"], gr["total_gst_amount"]])
    for sec in report["canteens"]:
        for gr in sec["gst_summary"]:
            ws.append(["Canteen", sec["canteen"]["name"], gr["gst_rate"], gr["gross_amount"],
                       gr["taxable_amount"], gr["cgst_amount"], gr["sgst_amount"],
                       gr["total_gst_amount"]])
    if ws.max_row > 1:
        _money_cells(ws, 2, ws.max_row, [4, 5, 6, 7, 8])
    _autosize(ws, [13, 22, 12, 14, 15, 12, 12, 13])

    # ---------------- Payment Summary ----------------
    ws = wb.create_sheet("Payment Summary")
    _write_header(ws, ["Scope", "Canteen", "Payment Method", "Orders", "Amount"])
    for p in report["combined_payment_summary"]:
        ws.append(["All Canteens", "", p["payment_mode"], p["order_count"], p["gross_amount"]])
    for sec in report["canteens"]:
        for p in sec["payment_summary"]:
            ws.append(["Canteen", sec["canteen"]["name"], p["payment_mode"],
                       p["order_count"], p["gross_amount"]])
    if ws.max_row > 1:
        _money_cells(ws, 2, ws.max_row, [5])
    _autosize(ws, [13, 22, 20, 10, 14])

    # ---------------- Transactions ----------------
    ws = wb.create_sheet("Transactions")
    _write_header(ws, ["Canteen", "Order ID", "Time", "Customer", "Guest ID", "Item", "Qty",
                       "Unit Price", "GST %", "Gross", "Taxable", "CGST", "SGST", "Total GST",
                       "Payment", "Order Type", "Status"])
    for sec in report["canteens"]:
        for txn in sec["transactions"]:
            for item in txn["items"]:
                ws.append([
                    sec["canteen"]["name"], txn["order_id"], txn["time"], txn["customer"],
                    txn["guest_code"] or "", item["name"], item["quantity"], item["unit_price"],
                    item["gst_rate"], item["gross_amount"], item["taxable_amount"],
                    item["cgst_amount"], item["sgst_amount"], item["total_gst_amount"],
                    txn["payment_mode"], txn["order_type"], txn["status"],
                ])
    if ws.max_row > 1:
        _money_cells(ws, 2, ws.max_row, [8, 10, 11, 12, 13, 14])
    _autosize(ws, [20, 9, 8, 20, 12, 22, 6, 11, 8, 12, 12, 11, 11, 12, 10, 12, 11])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_daily_report_excel(report: dict) -> bytes:
    wb = Workbook()

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    t = report["totals"]

    ws["A1"] = "COLLEGE SAAPAADU - Daily Sales & GST Report"
    ws["A1"].font = Font(bold=True, size=13)

    meta = [
        ("College", report["college"]["name"]),
        ("Canteen", report["canteen"]["name"]),
        ("Report Date", _fmt_date(report["report_date"])),
        ("Generated At", report["generated_at"][:19].replace("T", " ") + " IST"),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=k).font = BOLD
        ws.cell(row=i, column=2, value=v)

    totals = [
        ("Total Orders", t["order_count"], None),
        ("Total Items Sold", t["item_count"], None),
        ("Gross Sales (GST inclusive)", t["gross_sales"], MONEY_FMT),
        ("Taxable Sales Value", t["taxable_sales"], MONEY_FMT),
        ("CGST", t["cgst_amount"], MONEY_FMT),
        ("SGST", t["sgst_amount"], MONEY_FMT),
        ("Total GST Collected", t["total_gst"], MONEY_FMT),
    ]
    ws.cell(row=8, column=1, value="DAILY TOTALS").font = BOLD
    for i, (k, v, fmt) in enumerate(totals, start=9):
        ws.cell(row=i, column=1, value=k)
        c = ws.cell(row=i, column=2, value=v)
        if fmt:
            c.number_format = fmt
        if k == "Total GST Collected":
            ws.cell(row=i, column=1).font = BOLD
            c.font = BOLD

    if report.get("contains_estimated_values"):
        ws.cell(row=18, column=1,
                value=("Note: some lines predate per-transaction tax capture and were "
                       "reconstructed from current menu pricing."))
    _autosize(ws, [32, 34])

    # ---------------- Transactions ----------------
    ws = wb.create_sheet("Transactions")
    _write_header(ws, ["S.No", "Order ID", "Time", "Customer", "Guest ID", "Item", "Qty",
                       "Unit Price", "GST %", "Gross", "Taxable", "CGST", "SGST", "Total GST",
                       "Payment", "Order Type", "Status"])
    r = 2
    n = 0
    for txn in report["transactions"]:
        for item in txn["items"]:
            n += 1
            ws.append([
                n, txn["order_id"], txn["time"], txn["customer"], txn["guest_code"] or "",
                item["name"], item["quantity"], item["unit_price"], item["gst_rate"],
                item["gross_amount"], item["taxable_amount"], item["cgst_amount"],
                item["sgst_amount"], item["total_gst_amount"],
                txn["payment_mode"], txn["order_type"], txn["status"],
            ])
            r += 1
    if r > 2:
        _money_cells(ws, 2, r - 1, [8, 10, 11, 12, 13, 14])
    _autosize(ws, [6, 9, 8, 22, 12, 24, 6, 11, 8, 12, 12, 11, 11, 12, 10, 12, 11])

    # ---------------- Item summary ----------------
    ws = wb.create_sheet("Item Summary")
    _write_header(ws, ["Item", "Total Qty", "GST %", "Gross Sales", "Taxable Value",
                       "CGST", "SGST", "Total GST"])
    for it in report["item_summary"]:
        ws.append([it["name"], it["quantity"], it["gst_rate"], it["gross_amount"],
                   it["taxable_amount"], it["cgst_amount"], it["sgst_amount"],
                   it["total_gst_amount"]])
    if ws.max_row > 1:
        _money_cells(ws, 2, ws.max_row, [4, 5, 6, 7, 8])
    _autosize(ws, [28, 11, 8, 14, 15, 12, 12, 13])

    # ---------------- GST summary ----------------
    ws = wb.create_sheet("GST Summary")
    _write_header(ws, ["GST Rate %", "Gross", "Taxable Value", "CGST", "SGST", "Total GST"])
    for g in report["gst_summary"]:
        ws.append([g["gst_rate"], g["gross_amount"], g["taxable_amount"],
                   g["cgst_amount"], g["sgst_amount"], g["total_gst_amount"]])
    if ws.max_row > 1:
        _money_cells(ws, 2, ws.max_row, [2, 3, 4, 5, 6])
    _autosize(ws, [12, 14, 15, 12, 12, 13])

    # ---------------- Payment summary ----------------
    ws = wb.create_sheet("Payment Summary")
    _write_header(ws, ["Payment Method", "Orders", "Amount"])
    total = 0.0
    for p in report["payment_summary"]:
        ws.append([p["payment_mode"], p["order_count"], p["gross_amount"]])
        total += float(p["gross_amount"])
    last = ws.max_row
    ws.append(["Total", "", round(total, 2)])
    for c in range(1, 4):
        ws.cell(row=ws.max_row, column=c).font = BOLD
    if last >= 2:
        _money_cells(ws, 2, ws.max_row, [3])
    else:
        ws.cell(row=ws.max_row, column=3).number_format = MONEY_FMT
    _autosize(ws, [22, 10, 14])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
